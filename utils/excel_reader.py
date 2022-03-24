import os.path

from openpyxl import load_workbook
from utils.log import Logger

# 加载存在的 excel 文件: 默认可读写
class ExcelReader():
    def __init__(self):
        self.cur_file = ""
        self.cur_sheet = ""
        self.log = Logger()
        self.basic_init()
        pass

    def basic_init(self):
        self.log.log_file_name = "excel_reader"

    def update_file_name(self, excel_file_name):
        self.cur_file = excel_file_name

    def update_sheet_name(self, sheet_name):
        self.cur_sheet = sheet_name

    def get_sheet_all_data(self):
        """
        根据Excel文件名,sheet名读取信息
        :return: 以矩阵的方式返回二维列表
        """
        try:
            wb = load_workbook(self.cur_file, read_only=True)
        except:
            self.log.error("文件不存在, " + self.cur_file)
            return None

        try:
            ws = wb[self.cur_sheet]
        except:
            self.log.error("所要读取的标签页不存在, " + self.cur_sheet)
            return None

        data = []
        for i in range(1, ws.max_row + 1):
            data.append([])
            for j in range(1, ws.max_column + 1):
                data[i - 1].append(ws.cell(row=i, column=j).value)
        return data

    def get_sheet_data_by_file_and_sheet_name(self, excel_file_name, sheet_name="Sheet1"):
        self.update_file_name(excel_file_name)
        self.update_sheet_name(sheet_name)
        return self.get_sheet_all_data()


    def get_sheet_basic_info(self):
        """
        :return: 最大行数,最大列数
        """
        pass

rd_excel = r"E:\代码空间\auto_test_framework\test_file\测试用例数据.xlsx"
rd_sheet = "Sheet1"
excel = ExcelReader()
data = excel.get_sheet_data_by_file_and_sheet_name(rd_excel, rd_sheet)
print("data", data)

    #功能规划
    #掺入文件名,返回该文件的基本信息,包括标签页,以及每个标签页的最大行数和最大列数
    #传入文件路径,标签页名称,返回






# 版权声明：本文为CSDN博主「魔小明」的原创文章，遵循CC 4.0 BY-SA版权协议，转载请附上原文出处链接及本声明。
# 原文链接：https://blog.csdn.net/David_Dai_1108/article/detaimd
# ls/78702032